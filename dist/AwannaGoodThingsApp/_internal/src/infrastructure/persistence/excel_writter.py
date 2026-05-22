"""
Adapter: ExcelWriter
"""

import pandas as pd
from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from domain.entities.sales_summary import RegionalSummary


class ExcelWriter:
    """Write data to Excel with professional formatting."""
    
    def __init__(self):
        self._header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self._header_font = Font(color="FFFFFF", bold=True, size=11)
        self._border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def write_regional_summary(
        self,
        summaries: List[RegionalSummary],
        output_path: Path,
        include_details: bool = False
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df = self._create_summary_dataframe(summaries)
            summary_df.to_excel(writer, sheet_name='Ringkasan Wilayah', index=False)
            self._format_worksheet(writer.sheets['Ringkasan Wilayah'], summary_df)
            
            if include_details:
                details_df = self._create_details_dataframe(summaries)
                details_df.to_excel(writer, sheet_name='Detail Transaksi', index=False)
                self._format_worksheet(writer.sheets['Detail Transaksi'], details_df)
        
        return output_path
    
    def write_full_report(
        self,
        summaries: List[RegionalSummary],
        statistics: dict,
        output_path: Path
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            exec_df = self._create_executive_summary(statistics)
            exec_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            summary_df = self._create_summary_dataframe(summaries)
            summary_df.to_excel(writer, sheet_name='Ringkasan Wilayah', index=False)
            self._format_worksheet(writer.sheets['Ringkasan Wilayah'], summary_df)
            
            details_df = self._create_details_dataframe(summaries)
            details_df.to_excel(writer, sheet_name='Semua Transaksi', index=False)
            self._format_worksheet(writer.sheets['Semua Transaksi'], details_df)
        
        return output_path
    
    def _create_summary_dataframe(self, summaries: List[RegionalSummary]) -> pd.DataFrame:
        data = []
        for s in summaries:
            data.append({
                'Wilayah': s.region,
                'Total Transaksi': s.total_transactions,
                'Transaksi Paid': s.paid_transactions,
                'Transaksi Pending': s.pending_transactions,
                'Rate Pembayaran (%)': round(s.payment_percentage, 2),
                'Total Amount': s.total_amount,
                'Paid Amount': s.paid_amount
            })
        return pd.DataFrame(data)
    
    def _create_details_dataframe(self, summaries: List[RegionalSummary]) -> pd.DataFrame:
        data = []
        for summary in summaries:
            for t in summary.transactions:
                data.append({
                    'ID Transaksi': t.transaction_id,
                    'Tanggal': t.date,
                    'Produk': t.product_name,
                    'Wilayah': t.region,
                    'Status': t.payment_status,
                    'Amount': t.amount,
                    'Quantity': t.quantity
                })
        return pd.DataFrame(data)
    
    def _create_executive_summary(self, statistics: dict) -> pd.DataFrame:
        data = [
            {'Metrik': 'Total Wilayah', 'Nilai': statistics.get('total_regions', 0)},
            {'Metrik': 'Total Transaksi', 'Nilai': statistics.get('total_transactions', 0)},
            {'Metrik': 'Total Paid', 'Nilai': statistics.get('total_paid', 0)},
            {'Metrik': 'Total Pending', 'Nilai': statistics.get('total_pending', 0)},
            {'Metrik': 'Rate Pembayaran (%)', 'Nilai': round(statistics.get('overall_payment_rate', 0), 2)},
            {'Metrik': 'Total Amount', 'Nilai': statistics.get('total_amount', 0)},
            {'Metrik': 'Wilayah Teratas', 'Nilai': statistics.get('top_region', '-')}
        ]
        return pd.DataFrame(data)
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        for cell in worksheet[1]:
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._border
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = self._border
                cell.alignment = Alignment(vertical='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width