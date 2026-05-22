"""
Komponen: ExcelExporter
Export regional summary horizontal layout.
"""

from tkinter import filedialog, messagebox


class ExcelExporter:

    def __init__(self, get_data_fn):
        self.get_data = get_data_fn

    def save(self, default_filename="regional_summary"):

        dates_list, current_results = self.get_data()

        if not current_results:
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{default_filename}.xlsx"
        )

        if not file_path:
            return

        try:

            import pandas as pd

            from openpyxl import Workbook

            from openpyxl.styles import (
                Font,
                PatternFill,
                Alignment,
                Border,
                Side
            )

            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import FormulaRule

            # ======================================================
            # KUMPULKAN SEMUA PROVINSI
            # ======================================================

            all_provinces = set()

            normalized_data = {}

            for idx, date_str in enumerate(dates_list):

                result = current_results[date_str]

                df = result["dataframe"].copy()

                province_col = result["province_col"]

                rename_map = {}

                for col in df.columns:

                    lower = str(col).lower()

                    if "lead" in lower:
                        rename_map[col] = "leads"

                    elif "paid" in lower and "ratio" not in lower:
                        rename_map[col] = "paid"

                    elif col == province_col:
                        rename_map[col] = "province"

                df = df.rename(columns=rename_map)

                # ==========================================
                # PAID RATIO
                # ==========================================

                df["paid ratio"] = (
                    (
                        df["paid"].fillna(0)
                        /
                        df["leads"].replace(0, 1)
                    ) * 100
                ).round(1)

                df = df[
                    [
                        "province",
                        "leads",
                        "paid ratio",
                        "paid"
                    ]
                ]

                normalized_data[date_str] = df

                all_provinces.update(df["province"].tolist())

            all_provinces = sorted(all_provinces)

            # ======================================================
            # CREATE WORKBOOK
            # ======================================================

            wb = Workbook()

            ws = wb.active
            ws.title = "Regional Summary"

            # ======================================================
            # STYLING
            # ======================================================

            header_fill = PatternFill(
                start_color="3E6A9E",
                end_color="3E6A9E",
                fill_type="solid"
            )

            header_font = Font(
                color="FFFFFF",
                bold=True,
                size=11
            )

            green_fill = PatternFill(
                start_color="C6EFCE",
                end_color="C6EFCE",
                fill_type="solid"
            )

            red_fill = PatternFill(
                start_color="FFC7CE",
                end_color="FFC7CE",
                fill_type="solid"
            )

            yellow_fill = PatternFill(
                start_color="FFF2CC",
                end_color="FFF2CC",
                fill_type="solid"
            )

            bold_font = Font(
                bold=True
            )

            thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            center = Alignment(
                horizontal="center",
                vertical="center"
            )

            left_align = Alignment(
                horizontal="left",
                vertical="center"
            )

            # ======================================================
            # HEADER ROW 1
            # ======================================================

            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=2,
                end_column=1
            )

            province_cell = ws.cell(1, 1)
            province_cell.value = "province"

            current_col = 2

            for idx, date_str in enumerate(dates_list):

                start_col = current_col
                end_col = current_col + 2

                ws.merge_cells(
                    start_row=1,
                    start_column=start_col,
                    end_row=1,
                    end_column=end_col
                )

                ws.cell(
                    1,
                    start_col
                ).value = date_str

                current_col += 3

            # ======================================================
            # HEADER ROW 2
            # ======================================================

            current_col = 2

            for _ in dates_list:

                ws.cell(2, current_col).value = "leads"
                ws.cell(2, current_col + 1).value = "paid ratio"
                ws.cell(2, current_col + 2).value = "paid"

                current_col += 3

            # ======================================================
            # STYLE HEADER
            # ======================================================

            for row in ws.iter_rows(min_row=1, max_row=2):

                for cell in row:

                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin
                    cell.alignment = center

            # ======================================================
            # DATA UTAMA
            # ======================================================

            row_idx = 3

            for province in all_provinces:

                ws.cell(row_idx, 1).value = province

                current_col = 2

                for date_str in dates_list:

                    df = normalized_data[date_str]

                    province_row = df[
                        df["province"] == province
                    ]

                    if not province_row.empty:

                        leads = province_row.iloc[0]["leads"]
                        ratio = province_row.iloc[0]["paid ratio"]
                        paid = province_row.iloc[0]["paid"]

                    else:

                        leads = 0
                        ratio = 0
                        paid = 0

                    ws.cell(row_idx, current_col).value = leads
                    ws.cell(row_idx, current_col + 1).value = ratio / 100
                    ws.cell(row_idx, current_col + 2).value = paid

                    current_col += 3

                row_idx += 1

            # ======================================================
            # FORMAT PERSENTASE
            # ======================================================

            for col in range(1, ws.max_column + 1):

                if ws.cell(2, col).value == "paid ratio":

                    for row in range(3, 37):

                        ws.cell(
                            row,
                            col
                        ).number_format = '0.0%'

            # ======================================================
            # BARIS PETUNJUK 37
            # ======================================================

            last_col = ws.max_column
            last_col_letter = get_column_letter(last_col)

            ws.merge_cells(
                f"A37:{last_col_letter}37"
            )

            ws["A37"] = "Tempelkan salinan data dari spreadsheet di kolom A38"

            ws["A37"].fill = yellow_fill
            ws["A37"].font = bold_font
            ws["A37"].alignment = left_align

            # ======================================================
            # AREA SALINAN (38 - 71)
            # ======================================================

            for row in range(38, 72):

                for col in range(1, ws.max_column + 1):

                    ws.cell(row, col).border = thin
                    ws.cell(row, col).alignment = center

            # ======================================================
            # BARIS PETUNJUK 72
            # ======================================================

            ws.merge_cells(
                f"A72:{last_col_letter}72"
            )

            ws["A72"] = "Hasil pemeriksaan berada di baris 73"

            ws["A72"].fill = yellow_fill
            ws["A72"].font = bold_font
            ws["A72"].alignment = left_align

            # ======================================================
            # HASIL PEMERIKSAAN
            # ======================================================

            result_start_row = 73

            for idx, province in enumerate(all_provinces):

                result_row = result_start_row + idx
                source_row = 3 + idx
                compare_row = 38 + idx

                ws.cell(result_row, 1).value = province

                current_col = 2

                while current_col <= ws.max_column:

                    leads_col_letter = get_column_letter(current_col)
                    ratio_col_letter = get_column_letter(current_col + 1)
                    paid_col_letter = get_column_letter(current_col + 2)

                    # ======================================
                    # LEADS CHECK
                    # ======================================

                    leads_formula = (
                        f'=IF($A{compare_row}="","",'
                        f'IF({leads_col_letter}{source_row}-{leads_col_letter}{compare_row}=0,'
                        f'"Aman",'
                        f'IF({leads_col_letter}{source_row}-{leads_col_letter}{compare_row}<0,'
                        f'"Kelebihan "&ABS({leads_col_letter}{source_row}-{leads_col_letter}{compare_row}),'
                        f'"Kurang "&({leads_col_letter}{source_row}-{leads_col_letter}{compare_row})'
                        f')))'
                    )

                    ws.cell(
                        result_row,
                        current_col
                    ).value = leads_formula

                    # ======================================
                    # PAID RATIO
                    # ======================================

                    ws.cell(
                        result_row,
                        current_col + 1
                    ).value = "-"

                    # ======================================
                    # PAID CHECK
                    # ======================================

                    paid_formula = (
                        f'=IF($A{compare_row}="","",'
                        f'IF({paid_col_letter}{source_row}-{paid_col_letter}{compare_row}=0,'
                        f'"Aman",'
                        f'IF({paid_col_letter}{source_row}-{paid_col_letter}{compare_row}<0,'
                        f'"Kelebihan "&ABS({paid_col_letter}{source_row}-{paid_col_letter}{compare_row}),'
                        f'"Kurang "&({paid_col_letter}{source_row}-{paid_col_letter}{compare_row})'
                        f')))'
                    )

                    ws.cell(
                        result_row,
                        current_col + 2
                    ).value = paid_formula

                    current_col += 3

            # ======================================================
            # STYLE DATA
            # ======================================================

            for row in ws.iter_rows(
                min_row=3,
                max_row=106
            ):

                for cell in row:

                    cell.border = thin
                    cell.alignment = center

            # ======================================================
            # CONDITIONAL FORMATTING
            # ======================================================

            current_col = 2

            while current_col <= ws.max_column:

                col_letter = get_column_letter(current_col)

                range_text = f"{col_letter}73:{col_letter}106"

                ws.conditional_formatting.add(
                    range_text,
                    FormulaRule(
                        formula=[
                            f'ISNUMBER(SEARCH("Aman",{col_letter}73))'
                        ],
                        fill=green_fill
                    )
                )

                ws.conditional_formatting.add(
                    range_text,
                    FormulaRule(
                        formula=[
                            f'OR(ISNUMBER(SEARCH("Kelebihan",{col_letter}73)),ISNUMBER(SEARCH("Kurang",{col_letter}73)))'
                        ],
                        fill=red_fill
                    )
                )

                paid_col = current_col + 2
                paid_col_letter = get_column_letter(paid_col)

                paid_range = f"{paid_col_letter}73:{paid_col_letter}106"

                ws.conditional_formatting.add(
                    paid_range,
                    FormulaRule(
                        formula=[
                            f'ISNUMBER(SEARCH("Aman",{paid_col_letter}73))'
                        ],
                        fill=green_fill
                    )
                )

                ws.conditional_formatting.add(
                    paid_range,
                    FormulaRule(
                        formula=[
                            f'OR(ISNUMBER(SEARCH("Kelebihan",{paid_col_letter}73)),ISNUMBER(SEARCH("Kurang",{paid_col_letter}73)))'
                        ],
                        fill=red_fill
                    )
                )

                current_col += 3

            # ======================================================
            # COLUMN WIDTH
            # ======================================================

            ws.column_dimensions["A"].width = 35

            current_col = 2

            while current_col <= ws.max_column:

                ws.column_dimensions[
                    get_column_letter(current_col)
                ].width = 10

                ws.column_dimensions[
                    get_column_letter(current_col + 1)
                ].width = 10

                ws.column_dimensions[
                    get_column_letter(current_col + 2)
                ].width = 10

                current_col += 3

            # ======================================================
            # ROW HEIGHT
            # ======================================================

            ws.row_dimensions[37].height = 24
            ws.row_dimensions[72].height = 24

            # ======================================================
            # FREEZE
            # ======================================================

            ws.freeze_panes = "B3"

            # ======================================================
            # SAVE
            # ======================================================

            wb.save(file_path)

            messagebox.showinfo(
                "Sukses",
                f"File berhasil disimpan:\n{file_path}"
            )

        except Exception as e:

            messagebox.showerror(
                "Error",
                f"Gagal export:\n{str(e)}"
            )